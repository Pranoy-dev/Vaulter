"""Text extraction from PDF, DOCX, and XLSX files."""

from __future__ import annotations

import io


def extract_text(content: bytes, file_extension: str) -> str:
    """Extract text content from a file based on its extension.

    Returns extracted text or empty string if extraction fails.
    """
    ext = file_extension.lower().lstrip(".")
    try:
        if ext == "pdf":
            return _extract_pdf(content)
        elif ext == "docx":
            return _extract_docx(content)
        elif ext == "xlsx":
            return _extract_xlsx(content)
        elif ext == "pptx":
            return _extract_pptx(content)
        elif ext in ("txt", "csv", "eml"):
            return content.decode("utf-8", errors="replace")
    except Exception:
        return ""
    return ""


def _extract_pdf(content: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    parts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            parts.append(text)
    return "\n".join(parts)


def _extract_docx(content: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _extract_xlsx(content: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                parts.append(" ".join(cells))
    return "\n".join(parts)


def _extract_pptx(content: bytes) -> str:
    # Minimal PPTX extraction — parse XML directly to avoid heavy dependency
    import zipfile
    import xml.etree.ElementTree as ET

    parts = []
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as z:
            for name in sorted(z.namelist()):
                if name.startswith("ppt/slides/slide") and name.endswith(".xml"):
                    tree = ET.parse(z.open(name))
                    for elem in tree.iter():
                        if elem.text and elem.text.strip():
                            parts.append(elem.text.strip())
    except Exception:
        pass
    return "\n".join(parts)
